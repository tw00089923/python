import openpyxl as xlsx
import os ,datetime ,time
# 處理 xls 文件
#import xlrd 
#import xlwt 




class export_rt_wh:
    def __init__(self):
        self.path = os.path.abspath(".")
        self.pop_path = os.path.dirname(self.path)
        self.par_path ="N:\\IED\\pc-50\\Rouing 發行\\2017年"
        self.par_dir =[]
        self.suffix=".xlsx"
        self.date_year = self.load_datetime()
        self.index_G3 = self.loading_par()
        self.G3 = "RT-"+ self.date_year + self.index_G3 
        self.today = datetime.datetime.today()
    
    def load_datetime(self):
        return datetime.datetime.now().strftime("%y")
    def loading_par(self):
        try:
            return str(len(os.listdir(self.par_path)) + 1).zfill(3)
            print(self.par_dir)
        except:
            return str(0).zfill(3)

    def save_file(self):
        b = input("enter file name :")
        c = input("enter auth :")
        d = b + self.suffix
        e = "製作 :" + "2017" + "howard"
        wb = xlsx.load_workbook('DOC040 rouing.xlsx')
        ws = wb.active
        ws['G3'] = self.G3
        ws['C3'] = b 
        #ws.cell(27,5).value = 11
        wb.save(d)


   #wb = xlsx.load_workbook('DOC040 rouing.xlsx')
    #ws = wb.active
    #print(ws.max_row)
   # ws['G3']= "test"
    #suffix = ".xlsx"
  #  name = input("Enter name : ")
   # wb_name = name + suffix 

#wb.save(wb_name)
a = export_rt_wh()
a.save_file()
#print(b)

#wb.template = True
#wb.save('text.xlsx') 
# 
# 
# 

